# -*- coding: utf-8 -*-
"""
Created on Mon Dec  2 14:50:34 2019

@author: richa

Spyder Editor
This is a temporary script
"""
import random
import openpyxl
import matplotlib.pyplot as plt
#
# Define RMD_calc function - takes RMD schedules and assets, returns RMD
#
def RMD_calc(year, t_NQ_Assets):
    # schedule of RMD for RF, JR
    rmf_RMD_sched = [1000000,100000,27.4,26.5,25.6,24.7,23.8,22.9,22,21.2,20.3,19.5,
                  18.7,17.9,17.1,16.3,15.5,14.8,14.1,13.4,12.7,12,11.4,10.8,10.2,9.6]
    jr_RMD_sched = [1000000,100000,1000000,1000000,1000000,1000000,1000000,1000000,27.4
                    ,26.5,25.6,24.7,23.8,22.9,22,21.2,20.3,19.5,
                    18.7,17.9,17.1,16.3,15.5,14.8,14.1]
    rmf_RMD = t_NQ_Assets * (1 / rmf_RMD_sched[year - 1]) *.75 # adjust for RF assets ~ 75% of total
    jr_RMD = t_NQ_Assets * (1 /jr_RMD_sched[year -1]) *.25 # adjust for JR assets ~ 25% of total
    #print(rmf_RMD, jr_RMD)
    RMD = jr_RMD + rmf_RMD
    return RMD # end of RMD_calc
#
# workbook object is created from file "montecarloparams.xlsx" 
#
wb_obj = openpyxl.load_workbook("C:/Users/richa/.spyder-py3/montecarloparams.xlsx") 
# Get workbook active sheet object, default is first sheet
sheet_obj = wb_obj.active 
# Cell objects also have row, column, attributes. Note: The first row or  
# column integer is 1, not 0. Cell object is created by using  
# sheet object's cell obj and sheet_obj methods. 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
# Print title in cell A1 to verify  
print("Sheet title: ", cell_obj.value)
# 
# Monte Carlo model for investments - Models account depletion under a set
# of assumptions.Get input variables from spreadsheet
#
NQ_Assets = sheet_obj.cell(row = 7, column = 2).value
print ("NQ assets (1000): ", NQ_Assets)
#
NQ_Return = sheet_obj.cell (row = 3, column = 2).value # 40 year avg for bonds is .058
print ("NQ return (%): ", NQ_Return)
#
NQ_Return_Sigma = sheet_obj.cell (row = 4, column = 2).value
print ("NQ std dev (%): ", NQ_Return_Sigma)
#
Q_Assets = sheet_obj.cell (row = 8, column = 2).value
print ("Q assets: ", Q_Assets)
#
Q_Return = sheet_obj.cell (row = 5, column = 2).value # 40 year avg for stocks is 7%
print ("Q return (%): ", Q_Return)
#
Q_Return_Sigma = sheet_obj.cell (row = 6, column = 2).value # 40 year avg for stocks is 12%
print("Q std dev (%): ", Q_Return_Sigma)
#
years = sheet_obj.cell (row = 15, column = 2).value # How many years
print("Years: ", years)
#
cycles = sheet_obj.cell (row = 16, column = 2).value # how many model iterations
print ("Monte Carlo iterations: ", cycles)
#
Net_SS = sheet_obj.cell (row = 11, column = 2).value
print("Net Social Security: ", Net_SS)
#
Net_Annuities = sheet_obj.cell (row = 17, column = 2).value # NQ annuities
print("Net NQ Annuities: ", Net_Annuities)
#
RMD = sheet_obj.cell (row = 12, column = 2).value # Qual annuity plus additional $43K
print("RMD: ", RMD)
#
Net_RMD = sheet_obj.cell (row = 13, column = 2).value # conservative st - 15% avg tax rate
print("Net RMD: ", Net_RMD)
#
Annual_Required = sheet_obj.cell (row = 9, column = 2).value
print("Annual budget: ", Annual_Required)
#
SS_Cola = sheet_obj.cell (row = 14, column = 2).value
print("Soc Sec COLA: ", SS_Cola)
#
Inflation = sheet_obj.cell (row = 10, column = 2).value
print("Annual inflation: ", Inflation)
#
# Main logic loop - cycles is number of times to run N-year simulation,
# accumulating a new final asset value at the end of each cycle
#
Results = [] # empty list to hold results of each cycle for later
ZeroYear = [] # empty list to collect years that NQ assets are exhausted

for cycle in range (1, cycles+1):
    # setup/restore initial conditions for next Monte Carlo iteration
    t_NQ_Assets = NQ_Assets
    t_Net_SS = Net_SS
    t_Q_Assets = Q_Assets
    t_Annual_Required = Annual_Required
    zero_flag = False
    for year in range (1, years+1):
        #
        # Non qual assets
        #
        if (t_NQ_Assets > 0): # check to see if there aare any NQ assets left
            t_NQ_return = random.gauss(mu = NQ_Return, sigma = NQ_Return_Sigma) 
            NQ_Earnings = t_NQ_Assets * t_NQ_return *.95 #assumes  NQ$ are F tax free
            RMD = RMD_calc(year, t_NQ_Assets)
            Net_RMD = RMD * (1 - RMD_Tax_Rate) # adjust RMD for tax
            # print Net_RMD)
            Total_Income = NQ_Earnings + t_Net_SS + Net_Annuities + Net_RMD
            Shortfall = t_Annual_Required - Total_Income
            t_NQ_Assets = t_NQ_Assets - Shortfall
            NQ_Cache = t_NQ_Assets  # preserve this value so it can be used when NQ drops < 0
        else:
            # NQ assets have run out, need to correct for the fact that <0 is detected the
            # year after assets cross the zero line, so
            # must subtract last NQ assets from existing Q assets to get true value
                if (zero_flag == False): # first time that NQ < 0
                    t_Q_assets = t_Q_Assets + NQ_Cache # add last year (now a negative) to Qual assets
                    t_NQ_Assets = 0
                    ZeroYear.append (year - 1)
                    zero_flag = True # set this flag so subsequent passes just go to ELSE
                else:
                    t_NQ_Assets = 0 # not the first time through, this assignment really does nothing
        #
        # now qual assets - when NQ assets are depleted, required income is
        # taken from Q assets
        #
        t_Q_Return = random.gauss(mu = Q_Return, sigma = Q_Return_Sigma)       
        t_Q_Earnings = (t_Q_Assets * t_Q_Return)
        RMD = RMD_calc(year, t_NQ_Assets)
        Net_RMD = RMD * (1 - RMD_Tax_Rate) # adjust for tax
        # now see if we are taaking required income out of Qual funds
        if (t_NQ_Assets == 0): # take additional Q expenses to compensate for zero NQ assets
            Total_Income = t_Q_Earnings + t_Net_SS + Net_Annuities + Net_RMD
            Shortfall = t_Annual_Required - Total_Income
            t_Q_Assets = t_Q_Assets - Shortfall - RMD
        else:  # still have NQ assets, Qual assets continue to accumulate ex RMD 
            t_Q_Assets = t_Q_Assets + t_Q_Earnings - RMD
        #
        # print ("Yr:%2d R:%5.3f NQ:%5.2f TI: %5.2f TNQ: %6.1f TQ: %6.1f" % (year, t_NQ_return, NQ_Earnings, Total_Income, t_NQ_Assets, t_Q_Assets))
        # increment annual requirements, COLA adjustments, etc
        #
        t_Annual_Required = t_Annual_Required * (1 + Inflation)
        t_Net_SS = t_Net_SS * (1 + SS_Cola)
  # grab total assets at end of each cycle and accumulate                
    Tot_Assets = t_Q_Assets + t_NQ_Assets #tot = Q + NQ
    Results.append(Tot_Assets)
#
print ("Completed ", cycles, "Monte Carlo Simulations")
#
# plot results
#
num_bins = 100
plt.grid(True)
plt.legend(loc='best')
n, bins, patches = plt.hist(Results, num_bins, facecolor='blue', alpha=0.5, label= "Frequency")
plt.xlabel("Assets At End Of Simulation (1000s)")
plt.legend()
plt.show()
n, bins, patches = plt.hist(ZeroYear, num_bins, facecolor='blue', alpha=0.5, label= "Frequency")
plt.xlabel("Year NQ Assets Exhausted")
plt.grid(True)
plt.legend()
plt.show()