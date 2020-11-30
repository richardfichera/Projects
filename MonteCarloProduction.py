#!/usr/bin/env python3.8
# Import necessary modules
import random
from tkinter import *
import openpyxl
import matplotlib.pyplot as plt
import numpy as np  # for use in accumulating and displaying annual averages

def annual_analysis (cycles, years, all_results):
    # print("In annual analysis routine")
    if cycles < 11: # following code is debug code, manually input small years/cycles, check results
        print(all_results)
        print('Mean by year: ', np.mean(all_results, axis=0))
        annual_array = np.mean(all_results, axis=0) # 1D array of means for each year
        annual_std = np.std(all_results, axis=0)
        print('Annual means in array: ', annual_array)
        print('Std in array: ', annual_std)
        print ('All results array shape: ', all_results.shape)  # validation of rows (cycles) and cols (years)
    annual_array = np.mean(all_results, axis=0)  # mean of the annual results columns by year in a 1D np array
    annual_std = np.std(all_results, axis=0)  # std deviation
    print('All results array shape: ', all_results.shape)  # validation of rows (cycles) and cols (years)
    #
    # plot results - avg annual asset balance, +- 1 std dev
    #
    plt.grid(True)
    plt.style.use('seaborn-dark')
    plt.plot(annual_array + annual_std, color='blue', label='+1 sigma')
    plt.plot(annual_array, color='black', label='Avg. Annual Balance')
    plt.plot(annual_array - annual_std, color='green', label='-1 sigma')
    plt.ylabel('Avg balance EOY')
    plt.legend()
    plt.show()

def RMD_calc(year, t_NQ_Assets):
    '''
    Calculate RMD for each year of the model

    Not real elegant - RMD for each person is hard coded into a list with an assumed start date of 2020 and first MRD
    in year each one turns 72.5, using currently published factors. RMD is determined by calculating total assets
     * 1/RMD factor for each year. Non-paying years are indicated by very large (10**6) factor, which will generate an
     error for nominal zero-pay years in the order of single digit $.
    '''
    # schedule of RMD for RF, JR, starting at 72, using 2019 tables
    # divides Q assets by RMD factor, 1000000 is used for years that have no RMD
    rmf_RMD_sched = [0, 0, 27.4, 26.5, 25.6, 24.7, 23.8, 22.9, 22, 21.2, 20.3, 19.5,
                     18.7, 17.9, 17.1, 16.3, 15.5, 14.8, 14.1, 13.4, 12.7, 12, 11.4, 10.8, 10.2, 9.6]
    jr_RMD_sched = [0, 0, 0, 0, 0, 0, 0, 0, 27.4
        , 26.5, 25.6, 24.7, 23.8, 22.9, 22, 21.2, 20.3, 19.5,
                    18.7, 17.9, 17.1, 16.3, 15.5, 14.8, 14.1]
    if rmf_RMD_sched[year-1] == 0:
        rmf_RMD = 0.0
    else:
        rmf_RMD = t_NQ_Assets * (1 / rmf_RMD_sched[year-1]) * .75  # adjust for RF assets ~ 75% of total
    if jr_RMD_sched[year-1] == 0:
        jr_RMD = 0.0
    else:
        jr_RMD = t_NQ_Assets * (1 / jr_RMD_sched[year-1]) * .25  # adjust for JR assets ~ 25% of total
    # print(rmf_RMD, jr_RMD)
    RMD = jr_RMD + rmf_RMD
    return RMD  # end of RMD_calc


def xls_get_input():
    # workbook object is created from file C:\Users\richa\Downloads\montecarloparams.xlsx
    wb_obj = openpyxl.load_workbook("C:/Users/richa/Downloads/montecarloparams.xlsx")
    # Get workbook active sheet object, default is first sheet
    sheet_obj = wb_obj.active
    # Cell objects also have row, column, attributes. Note: The first row or
    # column integer is 1, not 0. Cell object is created by using
    # sheet object's cell obj and sheet_obj methods.
    cell_obj = sheet_obj.cell(row=1, column=1)
    # Print title in cell A1 to verify
    print("Sheet title: ", cell_obj.value)
    #
    # Input for Monte Carlo investment model - models account depletion under a set
    # of assumptions.Get input variables from spreadsheet
    #
    NQ_Assets = sheet_obj.cell(row=7, column=2).value
    print("NQ assets (1000): ", NQ_Assets)
    #
    NQ_Return = sheet_obj.cell(row=3, column=2).value  # 40 year avg for bonds is .058
    print("NQ return (%): ", NQ_Return)
    #
    NQ_Return_Sigma = sheet_obj.cell(row=4, column=2).value
    print("NQ std dev (%): ", NQ_Return_Sigma)
    #
    Q_Assets = sheet_obj.cell(row=8, column=2).value
    print("Q assets: ", Q_Assets)
    #
    Q_Return = sheet_obj.cell(row=5, column=2).value  # 40 year avg for stocks is 7%
    print("Q return (%): ", Q_Return)
    #
    Q_Return_Sigma = sheet_obj.cell(row=6, column=2).value  # 40 year avg for stocks is 12%
    print("Q std dev (%): ", Q_Return_Sigma)
    #
    years = sheet_obj.cell(row=15, column=2).value  # How many years
    print("Years: ", years)
    #
    cycles = sheet_obj.cell(row=16, column=2).value  # how many model iterations
    print("Monte Carlo iterations: ", cycles)
    #
    Net_SS = sheet_obj.cell(row=11, column=2).value
    print("Net Social Security: ", Net_SS)
    #
    Net_Annuities = sheet_obj.cell(row=17, column=2).value  # NQ annuities
    print("Net NQ Annuities: ", Net_Annuities)
    #
    # RMD = sheet_obj.cell(row=12, column=2).value  # Qual annuity plus additional $43K
    # print("RMD: ", RMD) # Not passing RMD as parameter, fix spreadsheet later
    #
    Annual_Required = sheet_obj.cell(row=9, column=2).value
    print("Annual budget: ", Annual_Required)
    #
    SS_Cola = sheet_obj.cell(row=14, column=2).value
    print("Soc Sec COLA: ", SS_Cola)
    #
    Inflation = sheet_obj.cell(row=10, column=2).value
    print("Annual inflation: ", Inflation)
    #
    RMD_Tax_Rate = sheet_obj.cell(row=18, column=2).value
    print("RMD Tax Rate: ", RMD_Tax_Rate)
    return (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
            Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
            RMD_Tax_Rate)  # end of spreadsheet input


def std_get_input():
    '''
    Get inout variables, some hard-coded, most frequently changed from console
    '''

    Annual_Required = float(input("Desired annual after-tax income:"))
    NQ_Assets = float(input("NQ investment assets (1000s):"))
    NQ_Return = .058  # 40 year avg for bonds is .058,use this for whole NQ portfolio for conservative bias
    NQ_Return_Sigma = .06  # 40 year avg for bonds is 3% , double for mixed portfolio
    Q_Assets = float(input("Q investment assets (1000s):"))
    Q_Return = .07  # 40 year avg for stocks is 7%
    Q_Return_Sigma = .12
    # 40 year avg for stocks is 12%
    years = int(input("How many years:"))  # How many years per cycle
    if years > 25:
        print('Maximum 25 years, setting to 25 years')
        years = 25
    cycles = int(input("How many Monte Carlo iterations?:"))  # how many model samples
    Net_SS = 39
    Net_Annuities = 9.8  # NQ annuities
    SS_Cola = .01
    Inflation = .02
    RMD_Tax_Rate = .15
    # return all vars here...
    return (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
            Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
            RMD_Tax_Rate)  # end of console + hard-coded input


def tkinter_window_test():
    '''
    GUI input module using Tkinter

    Built using GRID placement and SLIDER objects. Window priority tweaked to force window to top, then relinquished
    to enable results graphic display to have priority
    '''
    # print(tkinter_window_test.__doc__)
    root = Tk()
    root.attributes('-topmost', True)
    title = Label(root, text='Monte Carlo Simulation Data Entry', font=("Helvetica", 16), bg='#909', fg='white',
                  height=2, padx=100, pady=10)
    title.grid(row=0, column=0, columnspan=3)

    Annual_Req = Scale(activebackground='green', relief=GROOVE, length=200, from_=90, to=250, label='Annual Required',
                       orient=HORIZONTAL, bg='#CCC')
    Annual_Req.set(168)
    Annual_Req.grid(row=1)

    NQ_Ass = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=900, to=2000, label='NQ Assets',
                   orient=HORIZONTAL, bg='#CCC')
    NQ_Ass.set(1500)
    NQ_Ass.grid(row=2, column=0)

    Q_Ass = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=900, to=2000, label='Q Assets',
                  orient=HORIZONTAL, bg='#CCC')
    Q_Ass.set(1500)
    Q_Ass.grid(row=3, column=0)

    yrs = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=20, to=30, label='Years',
                orient=HORIZONTAL, bg='#CCC')
    yrs.set(25)
    yrs.grid(row=1, column=1)

    cyc = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=1000, to=10000, label='Model cycles',
                orient=HORIZONTAL, bg='#CCC')
    cyc.set(1000)
    cyc.grid(row=1, column=2)

    NQ_Ret = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=1.0, to=10.0,
                   label='NQ Return (%)', orient=HORIZONTAL, bg='#CCC')
    NQ_Ret.set(5.5)
    NQ_Ret.grid(row=2, column=1)

    NQ_Ret_Sig = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=1.0, to=10.0,
                       label='NQ Return Sigma (%)', orient=HORIZONTAL, bg='#CCC')
    NQ_Ret_Sig.set(6.0)
    NQ_Ret_Sig.grid(row=2, column=2)

    Q_Ret = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=1.0, to=10.0,
                  label='Q Return (%)', orient=HORIZONTAL, bg='#CCC')
    Q_Ret.set(6.7)
    Q_Ret.grid(row=3, column=1)

    Q_Ret_Sig = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=9.0, to=20.0,
                      label='Q Return Sigma (%)', orient=HORIZONTAL, bg='#CCC')
    Q_Ret_Sig.set(12.0)
    Q_Ret_Sig.grid(row=3, column=2)

    N_SS = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=30, to=40,
                 label='Net Social Security (1000s)', orient=HORIZONTAL, bg='#CCC')
    N_SS.set(39)
    N_SS.grid(row=4)

    N_Annuit = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=0, to=12,
                     label='Net NQ Annuities (1000s)', orient=HORIZONTAL, bg='#CCC')
    N_Annuit.set(10)
    N_Annuit.grid(row=4, column=1)

    RMD_Tax = Scale(root, activebackground='green', relief=GROOVE, length=200, from_=10, to=20,
                    label='RMD Tax Rate (%)', orient=HORIZONTAL, bg='#CCC')
    RMD_Tax.set(15)
    RMD_Tax.grid(row=4, column=2)

    Infl = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=1.0, to=10.0,
                 label='Inflation (%)', orient=HORIZONTAL, bg='#CCC')
    Infl.set(2.0)
    Infl.grid(row=5, column=1)

    S_Cola = Scale(root, resolution=.1, activebackground='green', relief=GROOVE, length=200, from_=.5, to=2.0,
                   label='SS COLA (%)', orient=HORIZONTAL, bg='#CCC')
    S_Cola.set(1.0)
    S_Cola.grid(row=5, column=0)

    Button(root, text='Submit', command=root.quit, bg='#CCC', activebackground='green').grid(row=7, column=1)
    mainloop()

    SS_Cola = float(S_Cola.get()) / 100.0
    print(SS_Cola)  # following print statements are for debug/confirmation -nobody seems to need them
    RMD_Tax_Rate = float(RMD_Tax.get()) / 100.0
    print(RMD_Tax_Rate)
    Annual_Required = float(Annual_Req.get())
    print(Annual_Required)  # print statements here for debug
    NQ_Assets = float(NQ_Ass.get())
    print(NQ_Assets)
    NQ_Return = float(NQ_Ret.get()) / 100.0
    print(NQ_Return)
    NQ_Return_Sigma = float(NQ_Ret_Sig.get()) / 100.0
    print(NQ_Return_Sigma)
    Q_Assets = float(Q_Ass.get())
    print(Q_Assets)
    Q_Return = float(Q_Ret.get()) / 100.0
    print(Q_Return)
    Q_Return_Sigma = float(Q_Ret_Sig.get()) / 100.0
    print(Q_Return_Sigma)
    years = int(yrs.get())
    print(years)
    cycles = int(cyc.get())
    print(cycles)
    Net_SS = float(N_SS.get())
    Net_Annuities = float(N_Annuit.get())
    print(Net_Annuities)
    Inflation = float(Infl.get()) / 100.0
    # print(type(Q_Return))
    # print('Scaled Q_Return =', Q_Return)
    # print(type(NQ_Return_Sigma))
    # print('Scaled NQ_Return_Sigma', NQ_Return_Sigma)
    # print(type(Q_Return_Sigma))
    # print('Scaled NQ Return', NQ_Return)
    # print(type(Inflation))
    # print('Scaled inflation = ', Inflation)
    root.attributes('-topmost', False)
    return (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
            Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
            RMD_Tax_Rate)


def main():
    '''
    Computes and displays Monte Carlo simulation of portfolio.

    Portfolio consists of two portions - Qualified (Q) and Non_Qualified (NQ), with NQ being drawn down first.
    Currently set up for two people with RMDs for NQ starting in year turned 72.5. NQ schedule and percentages
    are hard-coded. Program allows input of major variables for simulation:

    Required annual income
    Years per model cycle and number of iterations.
    Q and NQ starting assets along with returns and sigma for returns
    Soc Sec payments, COLA, RMD tax rate, Inflation

    Current handling of Q tax rates is not consistent, assumes taxes included in annual requirement but nets out
    fed tax from RMD distributions.

    Output shows distribution of final total portfolio value and distribution of year in which NQ assets are depleted.

    At runtime, users chooses input method: G-GUI, C-Console, X - Excel file in downloads directory.
    '''
    Results = []  # empty list to hold results of each cycle for later
    ZeroYear = []  # empty list to collect years that NQ assets are exhausted
    Annual_Req_Adjust = [1.0, 1.0, 1.0, 1.0, 1.0, .95, .95, .95, .95, .95, .90, .90, .90, .90, .90,
                         .90, .90, .90, .90, .90, 1.0, 1.0, 1.0, 1.0,
                         1.0]  # adjust required spend for age, then for extra help
    Param_Source = input("Use GUI(G), XLS(X) or Console(C) for input:").upper()
    if Param_Source == "X":
        (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
         Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
         RMD_Tax_Rate) = xls_get_input()
    elif Param_Source == 'C':
        (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
         Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
         RMD_Tax_Rate) = std_get_input()
    elif Param_Source == 'G':
        (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
         Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
         RMD_Tax_Rate) = tkinter_window_test()
    else:  # input error, default to console
        print('Input source error, defaulting to console input')
        (Annual_Required, NQ_Assets, NQ_Return, NQ_Return_Sigma, Q_Assets,
         Q_Return, Q_Return_Sigma, years, cycles, Net_SS, Net_Annuities, SS_Cola, Inflation,
         RMD_Tax_Rate) = std_get_input()
    #
    # Main logic loop - cycles is number of times to run N-year simulation,
    # accumulating a new final asset value at the end of each cycle
    #
    # first define numpy array to store results by cycle and year
    # results = 2D (years, cycles) numpy array, each year in 1 col, each row is 1 cycle
    #
    # all_results = np.arange(cycles * years).reshape(cycles , years)
    #
    all_results = np.zeros((cycles,years))

    for cycle in range(1, cycles+1):
        # setup/restore initial conditions for next Monte Carlo iteration
        t_NQ_Assets = NQ_Assets
        t_Net_SS = Net_SS
        t_Q_Assets = Q_Assets
        t_Annual_Required = Annual_Required
        zero_flag = False
        for year in range(1, years+1):
            #
            # Deplete non-qualified assets first
            #
            if (t_NQ_Assets > 0):  # check to see if there are any NQ assets left
                t_NQ_return = random.gauss(mu=NQ_Return, sigma=NQ_Return_Sigma)
                NQ_Earnings = t_NQ_Assets * t_NQ_return * .95  # assumes  NQ$ are F tax free
                RMD = RMD_calc(year, t_NQ_Assets)
                Net_RMD = RMD * (1 - RMD_Tax_Rate)  # adjust RMD for tax
                # print Net_RMD)
                Total_Income = NQ_Earnings + t_Net_SS + Net_Annuities + Net_RMD
                Shortfall = t_Annual_Required - Total_Income
                t_NQ_Assets = t_NQ_Assets - Shortfall
                NQ_Cache = t_NQ_Assets  # preserve this value so it can be used when NQ drops < 0
            else:
                # NQ assets have run out, need to correct for the fact that <0 is detected the
                # year after assets cross the zero line, so
                # must subtract last NQ assets from existing Q assets to get true value
                if (zero_flag == False):  # first time that NQ < 0
                    t_Q_assets = t_Q_Assets + NQ_Cache  # add last year (now a negative) to Qual assets
                    t_NQ_Assets = 0
                    ZeroYear.append(year - 1) # use year-1 because it ran out previous year
                    zero_flag = True  # set this flag so subsequent passes just go to ELSE
                else:
                    t_NQ_Assets = 0  # not the first time through, this assignment really does nothing
            #
            # now qual assets - when NQ assets are depleted, required income is
            # taken from Q assets
            #
            t_Q_Return = random.gauss(mu=Q_Return, sigma=Q_Return_Sigma)
            t_Q_Earnings = (t_Q_Assets * t_Q_Return)
            RMD = RMD_calc(year, t_Q_Assets)
            Net_RMD = RMD * (1 - RMD_Tax_Rate)  # adjust for tax
            # now see if we are taking required income out of Qual funds
            if (t_NQ_Assets == 0):  # take additional Q expenses to compensate for zero NQ assets
                Total_Income = t_Q_Earnings + t_Net_SS + Net_Annuities + Net_RMD
                Shortfall = (t_Annual_Required - Total_Income)/.85 #need to adjust for tax rate for NQ dist, assume .15
                t_Q_Assets = t_Q_Assets - Shortfall - RMD
            else:  # still have NQ assets, Qual assets continue to accumulate ex RMD
                t_Q_Assets = t_Q_Assets + t_Q_Earnings - RMD
            #
            # print ("Yr:%2d R:%5.3f NQ:%5.2f TI: %5.2f TNQ: %6.1f TQ: %6.1f" % (year, t_NQ_return, NQ_Earnings, Total_Income, t_NQ_Assets, t_Q_Assets))
            # increment annual requirements, COLA adjustments, etc
            #
            t_Annual_Required = t_Annual_Required * (1 + Inflation) * Annual_Req_Adjust[year - 1]
            t_Net_SS = t_Net_SS * (1 + SS_Cola)
            # write assets at end of each year into numpy results array
            yearly_total_assets = t_Q_Assets + t_NQ_Assets
            all_results[cycle-1,year-1] = yearly_total_assets
            #
        # grab total assets at end of each cycle and accumulate
        Tot_Assets = t_Q_Assets + t_NQ_Assets  # tot = Q + NQ at end of this cycle
        Results.append(Tot_Assets)
        #
    print("Completed ", cycle, "Monte Carlo Simulations")
    annual_display = input("Display annual average balances?").upper  # ask re annual balance display desired
    #
    # Print main histogram and year for depletion of NQ assets
    #
    num_bins = 100
    #plt.style.use('classic') # nice style - a little busy but good contrast
    plt.style.use('seaborn-dark') # nice and clean, 'fivethirtyeight' nice with line plots
    #plt.style.use('ggplot')
    # plt.grid(True)
    # create multiple plots via plt.subplots(rows,columns)
    fig, axes = plt.subplots(2)
    # one plot on each subplot
    axes[0].hist(Results, num_bins, facecolor='blue', alpha=0.5, label="Frequency")
    axes[1].hist(ZeroYear, num_bins, facecolor='blue', alpha=0.5, label="Frequency")
    axes[0].legend(['Final Value'])
    axes[1].legend(['Year NQ Assets Exhausted'])
    axes[0].grid(True)
    axes[1].grid(True)
    plt.show()
    if annual_display() == "Y":
        annual_analysis(cycles, years, all_results)

if __name__ == '__main__':
    main()
